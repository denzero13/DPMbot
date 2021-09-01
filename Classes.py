import pandas as pd
import openpyxl as op
from pymongo import MongoClient
from function import regex_kod, category, html_file_create, mail_send_message
from config import mongo_connect


class MoodleTestFormation:
    def __init__(self, path):
        self.path = path
        self.df = pd.read_csv(path)
        self.collum_name = [name for name in self.df]
        self.question = self.collum_name[8::3]
        self.response = self.collum_name[9::3]
        self.right_answer = self.collum_name[10::3]
        self.result_json = []
        self.emails = set()

    def preparation(self):
        for row_i in range(0, len(self.df)):
            for collum_i in range(0, int(len(self.question))):
                object_json = {self.collum_name[0]: self.df[self.collum_name[0]][row_i],
                               self.collum_name[1]: self.df[self.collum_name[1]][row_i],
                               self.collum_name[2]: self.df[self.collum_name[2]][row_i],
                               self.collum_name[3]: self.df[self.collum_name[3]][row_i],
                               self.collum_name[4]: self.df[self.collum_name[4]][row_i],
                               self.collum_name[5]: self.df[self.collum_name[5]][row_i],
                               self.collum_name[6]: self.df[self.collum_name[6]][row_i],
                               "Garde": self.df[self.collum_name[7]][row_i]}
                try:
                    kod = regex_kod(self.df[self.question[collum_i]][row_i])
                    qw = self.df[self.question[collum_i]][row_i][len(kod) + 1:].split(kod)
                except TypeError:
                    kod = "Kod not found"
                    qw = self.df[self.question[collum_i]][row_i]

                object_json["Kod"] = kod
                object_json["Category"] = category(kod[2])
                object_json["Question"] = qw[0]

                # User result
                object_json["Response"] = self.df[self.response[collum_i]][row_i]
                object_json["Right answer"] = self.df[self.right_answer[collum_i]][row_i]

                # Evaluation of answers
                if self.df[self.response[collum_i]][row_i] == self.df[self.right_answer[collum_i]][row_i]:
                    object_json["Correct answer"] = 1
                else:
                    object_json["Correct answer"] = 0
                # Ball inversion
                if self.df[self.response[collum_i]][row_i] == self.df[self.right_answer[collum_i]][row_i]:
                    object_json["Wrong answer"] = 0
                else:
                    object_json["Wrong answer"] = 1

                self.result_json.append(object_json)

    def mail_preparation(self):
        for row_i in range(0, len(self.df)):
            self.emails.add(self.df[self.collum_name[2]][row_i])

    def to_exel(self):
        self.preparation()
        df = pd.DataFrame(self.result_json)
        df.to_excel("documents/test_result.xlsx")

    def to_mongo(self, name):
        client = MongoClient(f"mongodb+srv://{mongo_connect}."
                             "hgdrv.mongodb.net/myFirstDatabase?retryWrites=true&w""=majority")
        db = client.Moodle
        collection = db[name]
        self.preparation()
        collection.insert_many(self.result_json)

    def to_mail(self):
        self.mail_preparation()
        for email in self.emails:
            html_file_create(email)

        mail_send_message(self.emails)


class Gift:
    def __init__(self, path):
        self.wb = op.load_workbook(path, data_only=True)  # Open exel file for work
        self.sheet = self.wb.active
        self.number_row = self.sheet.max_row  # Counts the number of rows
        self.main_category = 'Main category name'

        self.category = set()
        for i in range(2, self.number_row):
            self.category.add(self.sheet.cell(row=i, column=8).value)

    def data_formation(self):
        with open('documents/import.txt', 'tw', encoding='utf-8') as file:
            # Main Category
            file.write(f'// question: 0  name: Switch category to /top/{self.main_category}\n')
            file.write(f'$CATEGORY: /top/{self.main_category}'+'\n'*3)

            iteration = 1
            for teg in self.category:
                # Category
                file.write(f'// question: 0  name: Switch category to $cat1$/top/{self.main_category}/{teg}\n')
                file.write(f'$CATEGORY: /top/{self.main_category}/{teg}' + '\n' * 3)

                for i in range(2, self.number_row):

                    variant = {'A': '~', 'B': '~', 'C': '~', 'D': '~'}  # Answers value
                    true_answer = str(self.sheet.cell(row=i, column=6).value)

                    for v in variant.keys():  # Answer generation
                        if v in true_answer:
                            variant[v] = '='

                    if teg == self.sheet.cell(row=i, column=8).value:
                        # Question name
                        file.write(f'// question: {iteration} name: '+str(self.sheet.cell(row=i, column=1).value)+'\n')
                        # Kod in tag
                        file.write(f'// [tag:{teg}]' + '\n')
                        # Question title
                        file.write(f'::{teg}\t' + str(self.sheet.cell(row=i, column=1).value))
                        file.write(f'\:::[html]{teg}\t' + str(self.sheet.cell(row=i, column=1).value) + '\:{' + '\n')
                        # Answers
                        file.write('\t'+str(self.variant.get('A'))+str(self.sheet.cell(row=i, column=2).value)+'\n')
                        file.write('\t'+str(self.variant.get('B'))+str(self.sheet.cell(row=i, column=3).value)+'\n')
                        file.write('\t'+str(self.variant.get('C'))+str(self.sheet.cell(row=i, column=4).value)+'\n')
                        file.write('\t'+str(self.variant.get('D'))+str(self.sheet.cell(row=i, column=5).value)+'\n')
                        file.write('} '+'\n'*3)
                        iteration += 1