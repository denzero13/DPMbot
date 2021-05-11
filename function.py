import pandas as pd
import openpyxl as op
import re


def gift_generation(path):

    #Open exel file for work
    wb = op.load_workbook(path, data_only=True)
    sheet = wb.active

    #Counts the number of rows
    number_row = sheet.max_row

    main_category = 'Main category name'
    variant = {'A' : '~', 'B' : '~', 'C' : '~', 'D' : '~'}


    category = set()
    for i in range(2, number_row):
        category.add(sheet.cell(row=i, column=8).value)

    with open('documents/import.txt', 'tw', encoding='utf-8') as file:
        #Main Categori 
        file.write(f'// question: 0  name: Switch category to /top/{main_category}\n')
        file.write(f'$CATEGORY: /top/{main_category}'+'\n'*3)

        iteration = 1
        for cteg in category:
            #Categori
            file.write(f'// question: 0  name: Switch category to $cat1$/top/{main_category}/{cteg}\n')
            file.write(f'$CATEGORY: /top/{main_category}/{cteg}'+'\n'*3)

            for i in range(2, number_row):
                #Answer value
                variant = {'A' : '~', 'B' : '~', 'C' : '~', 'D' : '~'}
                true_answer = str(sheet.cell(row=i, column=6).value)

                #Answer genaration
                for v in variant.keys():
                     if v in true_answer:
                        variant[v] = '='  

                if cteg == sheet.cell(row=i, column=8).value:
                    #Qwestion name
                    file.write(f'// question: {iteration} name: '+str(sheet.cell(row=i, column=1).value)+'\n' )
                    #Kod in tag
                    file.write(f'// [tag:{cteg}]'+'\n' )
                    #Qwestion title
                    file.write(f'::{cteg}\t'+str(sheet.cell(row=i, column=1).value))
                    file.write(f'\:::[html]{cteg}\t'+str(sheet.cell(row=i, column=1).value)+'\:{'+'\n')
                    #Answers
                    file.write('\t'+str(variant.get('A'))+str(sheet.cell(row=i, column=2).value)+'\n')
                    file.write('\t'+str(variant.get('B'))+str(sheet.cell(row=i, column=3).value)+'\n')
                    file.write('\t'+str(variant.get('C'))+str(sheet.cell(row=i, column=4).value)+'\n')
                    file.write('\t'+str(variant.get('D'))+str(sheet.cell(row=i, column=5).value)+'\n')

                    file.write('} '+'\n'*3)
                    iteration += 1

        file.close()




def result_format(path):
    
    #Open exel file for work
    df = pd.read_excel(path, engine='openpyxl')
    
    #Generation collum name
    collum_name = []
    for i in df:
        collum_name.append(i)
    
    #Data slice
    qwestion = collum_name[8::3]
    response = collum_name[9::3]
    right_answer = collum_name[10::3]
    

    result_json = []
    for row_i in range(0, len(df)):
        for collum_i in range(0, int(len(qwestion))):

            object_json = {}
            
            #Information about a test proces
            object_json[collum_name[1]] = df[collum_name[1]][row_i]
            object_json[collum_name[2]] = df[collum_name[2]][row_i]
            object_json[collum_name[3]] = df[collum_name[3]][row_i]
            object_json[collum_name[4]] = df[collum_name[4]][row_i]
            object_json[collum_name[5]] = df[collum_name[5]][row_i]
            object_json[collum_name[6]] = df[collum_name[6]][row_i]
            object_json[collum_name[7]] = df[collum_name[7]][row_i]
            
            #Question identifiers
            try:
                kod = regex_kod(df[qwestion[collum_i]][row_i])
                qw = df[qwestion[collum_i]][row_i][len(kod)+1:].split(':')
            except TypeError:
                kod = 'Kod not found'
                qw = df[qwestion[collum_i]][row_i].split(':')

            object_json['Kod'] = kod
            object_json['Qwestion'] = qw[0]
           
            #User result
            object_json['Response'] = df[response[collum_i]][row_i]
            object_json['Right answer'] = df[right_answer[collum_i]][row_i]

            #Evaluation of answers
            if df[response[collum_i]][row_i] == df[right_answer[collum_i]][row_i]:
                object_json['Bal'] = 1
            else:
                object_json['Bal'] = 0

            result_json.append(object_json)
    
    #Create exel file         
    df = pd.DataFrame(result_json)
    df.to_excel('documents/test_result.xlsx')


def regex_kod(text):

    regex = r'^\d\S+'
    matches = re.finditer(regex, text, re.MULTILINE)
    
    for matchNum, match in enumerate(matches, start=1):
        return match.group()

